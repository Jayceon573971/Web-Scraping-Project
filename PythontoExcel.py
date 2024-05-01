import openpyxl as xl
from openpyxl.styles import Font

#create a new excel document
wb = xl.Workbook()

ws = wb.active #write to the one thats pop

#change title of ws
ws.title = 'First Sheet'

#create a new ws
wb.create_sheet(index=1, title= 'Second Sheet')

#write
ws['A1'] = 'Invoice'
fontobj = Font(name='Times New Roman',size=24,italic=False,bold=True)
ws['A1'].font = fontobj

ws['A2'] = 'Tires'
ws['A3'] = 'Brakes'
ws['A4'] = 'Alignment'

ws.merge_cells('A1:B1')
#ws.unmerge_cells('A1:B1')

ws['B2'] = 450
ws['B3'] = 225
ws['B4'] = 150

ws['A8'] = 'Total'
ws['A8'].font = Font(size=16,bold=True)
ws['B8'] = '=SUM(B2:B4)'


ws.column_dimensions['A'].width = 25

#Read the excel file - 'ProduceReport.xlsx'
#write to second sheet
#display the grandtotal of amt sold and total

write_sheet = wb['Second Sheet']
read_wb = xl.load_workbook('ProduceReport.xlsx')
read_ws = read_wb['ProduceReport']

for row in read_ws.iter_rows():
    ls = [i.value for i in row]
    write_sheet.append(ls)

#for i in read_ws:
#    for cell in i:
#        write_sheet[cell.coordinate] = cell.value

last_row = write_sheet.max_row

write_sheet.cell(last_row+2,3).value = 'AMT Sold Total'
write_sheet.cell(last_row+2,3).font = Font(size=16,bold=True)
write_sheet.cell(last_row+2,4).value = 'Total'
write_sheet.cell(last_row+2,4).font = Font(size=16,bold=True)
write_sheet.cell(last_row+3,3).value = '=SUM(C2:C41)'
write_sheet.cell(last_row+3,4).value = '=SUM(D2:D41)'

write_sheet.cell(last_row+4,3).value = 'AVG AMT Total'
write_sheet.cell(last_row+4,3).font = Font(size=16,bold=True)
write_sheet.cell(last_row+4,4).value = 'AVG Total'
write_sheet.cell(last_row+4,4).font = Font(size=16,bold=True)
write_sheet.cell(last_row+5,3).value = '=AVERAGE(C2:C41)'
write_sheet.cell(last_row+5,4).value = '=AVERAGE(D2:D41)'

write_sheet.column_dimensions['A'].width = 16
write_sheet.column_dimensions['B'].width = 16
write_sheet.column_dimensions['C'].width = 16
write_sheet.column_dimensions['D'].width = 16

for cell in write_sheet["C:C"]:
    cell.number_format = '#,##0'

for cell in write_sheet["D:D"]:
    cell.number_format = u'"$ "#,##0.00'

wb.save('PythontoExcel.xlsx')