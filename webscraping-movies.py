
from urllib.request import urlopen
from bs4 import BeautifulSoup
import openpyxl as xl
from openpyxl.styles import Font





#webpage = 'https://www.boxofficemojo.com/weekend/chart/'
webpage = 'https://www.boxofficemojo.com/year/2024/'

page = urlopen(webpage)			

soup = BeautifulSoup(page, 'html.parser')

title = soup.title

print(title.text)
##
wb = xl.Workbook()
ws = wb.active
ws.title = 'Box Office Report'

ws['A1'] = 'No.'
fontobj = Font(name='Times New Roman',size=24,italic=False,bold=True)
ws['A1'].font = fontobj
ws['B1'] = 'Movie Title'
fontobj = Font(name='Times New Roman',size=24,italic=False,bold=True)
ws['B1'].font = fontobj
ws['C1'] = 'Release Date'
fontobj = Font(name='Times New Roman',size=24,italic=False,bold=True)
ws['C1'].font = fontobj
ws['D1'] = 'Total Gross'
fontobj = Font(name='Times New Roman',size=24,italic=False,bold=True)
ws['D1'].font = fontobj
ws['E1'] = 'Theaters'
fontobj = Font(name='Times New Roman',size=24,italic=False,bold=True)
ws['E1'].font = fontobj
ws['F1'] = 'Average per theater'
fontobj = Font(name='Times New Roman',size=24,italic=False,bold=True)
ws['F1'].font = fontobj
table_rows = soup.findAll("tr")
##
for row_index, row in enumerate(table_rows[1:6], start=2):
    ws.cell(row=row_index, column=1).value = row_index - 1  # Numbering from 1 to 5
    td = row.findAll("td")
    Movie_Title = td[1].text
    Release_Date = td[8].text
    Total_Gross = int(td[7].text.replace(",","").replace("$",""))
    Theaters = int(td[6].text.replace(",",""))
    Theater_avg = Total_Gross / Theaters
    
    # Write data to respective cells
    ws.cell(row=row_index, column=2).value = Movie_Title
    ws.cell(row=row_index, column=3).value = Release_Date
    ws.cell(row=row_index, column=4).value = Total_Gross
    ws.cell(row=row_index, column=5).value = Theaters
    ws.cell(row=row_index, column=6).value = Theater_avg

##

##
ws.column_dimensions['F'].width = 16
ws.column_dimensions['B'].width = 16
ws.column_dimensions['C'].width = 16
ws.column_dimensions['E'].width = 16
ws.column_dimensions['D'].width = 16

for cell in ws["D:D"]:
    cell.number_format = u'"$ "#,##0.00'

for cell in ws["F:F"]:
    cell.number_format = u'"$ "#,##0.00'

wb.save('BoxOfficeReport.xlsx')